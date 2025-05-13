import io
import os
import json

import win32com
from PIL import Image, ImageGrab
import openpyxl
import xlwings as xw
import win32com.client as win32
import win32gui
import win32con
import time
from openpyxl.utils import get_column_letter


# -*- coding: utf-8 -*-
def listar_ventanas_office():
    office_windows = []

    def callback(hwnd, _):
        title = win32gui.GetWindowText(hwnd)
        class_name = win32gui.GetClassName(hwnd)
        if (win32gui.IsWindowVisible(hwnd) and title and
                ("Excel" in title or "Office" in title or class_name in ['NUIDialog', '#32770'])):
            office_windows.append((hwnd, title, class_name))

    win32gui.EnumWindows(callback, None)
    return office_windows


def cerrar_dialogos_office():
    dialogs_closed = 0
    windows = listar_ventanas_office()
    for hwnd, title, class_name in windows:
        try:
            if class_name in ['NUIDialog', '#32770'] and "Excel" not in title:
                win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
                time.sleep(2)
                if win32gui.IsWindow(hwnd):
                    win32gui.SendMessage(hwnd, win32con.WM_SYSCOMMAND, win32con.SC_CLOSE, 0)
                    time.sleep(1)
                if not win32gui.IsWindow(hwnd):
                    dialogs_closed += 1
        except:
            pass
    return dialogs_closed


class TSSProcessor:
    def __init__(self, config_path='config.json'):
        self.config = self._cargar_configuracion(config_path)
        self.data = {'textos': {}, 'imagenes': {}}
        self.capturas_dir = None
        self.nombre_sid = None

    # Configuración y helpers básicos

    def _cargar_configuracion(self, config_path):
        with open(config_path, encoding='utf-8') as f:
            return json.load(f)

    def _obtener_hoja_indice(self, workbook_type, sheet_name):
        return self.config['hojas'][workbook_type][sheet_name]

    def _generar_nombre_sid(self, tss_path):
        """Genera el nombre leyendo directamente del TSS según la configuración"""
        try:
            # 1. Cargar el archivo TSS
            wb_tss = openpyxl.load_workbook(tss_path, data_only=True)

            # 2. Obtener configuración de campos
            config_campos = self.config['nombre_sid']['campos']

            # 3. Leer valores directamente del TSS
            valores = {}
            for campo, config in config_campos.items():
                # Obtener hoja y celda desde la configuración
                sheet_index = self._obtener_hoja_indice('tss', config['hoja'])
                sheet = wb_tss.worksheets[sheet_index]
                celda = config['celda']

                # Leer valor y limpiar
                valor = str(sheet[celda].value or "").strip()
                valor = ''.join(c for c in valor if c not in '\\/:*?"<>|').replace(" ", "_")
                valores[campo] = valor or campo.upper()  # Si está vacío, usa el nombre del campo en mayúsculas

            # 4. Aplicar formato
            nombre_archivo = self.config['nombre_sid']['formato'].format(**valores)
            wb_tss.close()

            print(f"📄 Nombre generado: {nombre_archivo}")
            return nombre_archivo

        except Exception as e:
            print(f"⚠️ Error generando nombre: {str(e)}")
            # Nombre de respaldo con timestamp
            return f"SID_GENERADO_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"

    #Capturar informacion del tss

    def procesar_tss(self, tss_path, nombre_sid=None):
        """Procesa el TSS agrupando elementos por tipo para optimización"""
        print("\n=== PROCESANDO TSS ===")
        if nombre_sid is None:
            nombre_sid = self._generar_nombre_sid(tss_path)

        self.nombre_sid = nombre_sid
        self.capturas_dir = os.path.join("capturas", os.path.splitext(nombre_sid)[0])
        os.makedirs(self.capturas_dir, exist_ok=True)
        print(f"📁 Directorio de capturas creado: {self.capturas_dir}")

        wb_tss = openpyxl.load_workbook(tss_path, data_only=True)

        elementos_por_tipo = {
            'rango': [],
            'imagen': [],
            'texto': []
        }

        for elemento in self.config['elementos']:
            elementos_por_tipo[elemento['tipo']].append(elemento)

        # Paso 3: Procesar en orden optimizado
        # 3.1 Procesar rangos (requiere Excel COM)
        if elementos_por_tipo['rango']:
            self._procesar_rangos_agrupados(wb_tss, elementos_por_tipo['rango'])

        # 3.2 Procesar imágenes
        for elemento in elementos_por_tipo['imagen']:
            self._procesar_imagen(wb_tss, elemento)

        # 3.3 Procesar textos
        for elemento in elementos_por_tipo['texto']:
            self._procesar_texto(wb_tss, elemento)

        wb_tss.close()
        print("=== EXTRACCIÓN COMPLETADA ===")

    # Procesamiento interno del tss

    def _procesar_texto(self, wb_tss, elemento):
        sheet_index = self._obtener_hoja_indice('tss', elemento['origen']['hoja'])
        sheet = wb_tss.worksheets[sheet_index]
        valor = sheet[elemento['origen']['celda']].value
        self.data['textos'][elemento['nombre']] = str(valor).strip() if valor else ""
        print(f"Texto '{elemento['nombre']}' extraído: {self.data['textos'][elemento['nombre']]}")

    def _procesar_rangos_agrupados(self, wb_tss, elementos_rango):
        """Prepara el diccionario de rangos con información de hoja"""
        rangos_dict = {
            elem['nombre']: {
                'rango': elem['origen']['rango'],
                'hoja': elem['origen']['hoja']
            }
            for elem in elementos_rango
        }

        resultados = self.capturar_multiples_rangos(rangos_dict)

        for nombre, ruta_imagen in resultados.items():
            if ruta_imagen:
                self.data['imagenes'][nombre] = ruta_imagen

    def capturar_multiples_rangos(self, rangos_dict):
        """Captura múltiples rangos de hojas específicas y guarda en carpeta capturas"""
        excel = None
        resultados = {}

        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = True
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(os.path.abspath(self.tss_path))
            cerrar_dialogos_office()

            for nombre, config in rangos_dict.items():
                output_path = os.path.join(self.capturas_dir, f"{nombre}.png")

                # Obtener índice de hoja desde la configuración
                sheet_index = self._obtener_hoja_indice('tss', config['hoja']) + 1  # Excel usa base 1
                sheet = wb.Sheets(sheet_index)

                for intento in range(3):
                    try:
                        sheet.Range(config['rango']).CopyPicture(Appearance=1, Format=2)
                        time.sleep(2)

                        img = ImageGrab.grabclipboard()
                        if img:
                            img.save(output_path)
                            resultados[nombre] = output_path
                            print(f"✅ {nombre} guardado en {output_path} (Hoja: {config['hoja']})")
                            break
                    except Exception as e:
                        print(f"⚠️ Intento {intento + 1} para {nombre}: {str(e)}")
                        cerrar_dialogos_office()
                        time.sleep(1)
                else:
                    resultados[nombre] = None

            return resultados

        except Exception as e:
            print(f"❌ Error al capturar rangos: {str(e)}")
            return {}
        finally:
            if excel:
                try:
                    excel.DisplayAlerts = False
                    excel.Quit()
                except:
                    pass

    def _procesar_imagen(self, wb_tss, elemento):
        """Busca imágenes mostrando el rango de celdas de búsqueda"""
        try:
            sheet_index = self._obtener_hoja_indice('tss', elemento['origen']['hoja'])
            sheet = wb_tss.worksheets[sheet_index]
            celda = sheet[elemento['origen']['celda']]

            # Determinar coordenadas de búsqueda
            merged_range = self._encontrar_rango_combinado(celda, sheet)  # Pasar sheet como parámetro
            min_row, max_row, min_col, max_col = self._obtener_rango_expandido(celda, merged_range)

            # Convertir coordenadas numéricas a formato de letra de columna (A, B, C...)
            col_letter_start = openpyxl.utils.get_column_letter(min_col)
            col_letter_end = openpyxl.utils.get_column_letter(max_col)

            print(f"🔍 Buscando imagen {elemento['nombre']} en rango: "
                  f"{col_letter_start}{min_row}:{col_letter_end}{max_row} "
                  f"(Columnas {min_col}-{max_col}, Filas {min_row}-{max_row})")

            # Buscar imagen en el rango
            for img in sheet._images:
                img_top = img.anchor._from.row + 1
                img_left = img.anchor._from.col + 1

                if (min_row <= img_top <= max_row) and (min_col <= img_left <= max_col):
                    img_path = os.path.join(self.capturas_dir, f"{elemento['nombre']}.png")
                    image_bytes = img._data()
                    image = Image.open(io.BytesIO(image_bytes))
                    image.save(img_path)
                    self.data['imagenes'][elemento['nombre']] = img_path
                    print(f"✅ Imagen '{elemento['nombre']}' encontrada en posición: "
                          f"Columna {img_left}, Fila {img_top}")
                    return img_path

            print(f"⚠️ Imagen {elemento['nombre']} no encontrada en el rango especificado")
            return None

        except Exception as e:
            print(f"❌ Error al buscar imagen: {str(e)}")
            return None

    def _encontrar_rango_combinado(self, target_cell, sheet):
        """Encontrar rango combinado para la celda objetivo"""
        for merged_cell in sheet.merged_cells.ranges:  # Usar sheet en lugar de self.sheet_tss
            if target_cell.coordinate in merged_cell:
                print(f"\n ✅ Celda combinada encontrada: {merged_cell.coord}")
                return merged_cell
        print(f"\n ℹ️ Celda no está combinada")
        return None

    def _obtener_rango_expandido(self, target_cell, merged_range):
        """Obtener rango ampliado para búsqueda de imagen"""
        if merged_range:
            min_row, max_row = merged_range.min_row, merged_range.max_row
            min_col, max_col = merged_range.min_col, merged_range.max_col
        else:
            min_row = max_row = target_cell.row
            min_col = max_col = target_cell.column

        # Ampliar rango con márgenes
        return (
            max(1, min_row - 1),  # expanded_min_row
            max_row,              # expanded_max_row
            max(1, min_col - 1),  # expanded_min_col
            max_col               # expanded_max_col
        )

    #Generacion de sid

    def generar_sid(self, plantilla_path, output_path):
        """Genera el SID con los datos extraídos, soportando múltiples celdas destino"""
        print("\n=== GENERANDO SID ===")
        app = xw.App(visible=False)

        try:
            wb_sid = app.books.open(plantilla_path)

            # 1. Insertar textos (ahora soporta múltiples celdas destino)
            for elemento in self.config['elementos']:
                if elemento['tipo'] == 'texto' and elemento['nombre'] in self.data['textos']:
                    sheet_index = self._obtener_hoja_indice('sid', elemento['destino']['hoja'])
                    sheet = wb_sid.sheets[sheet_index]
                    valor = self.data['textos'][elemento['nombre']]

                    # Insertar el mismo valor en todas las celdas especificadas
                    for celda in elemento['destino']['celdas']:
                        sheet[celda].value = valor
                        print(f"Texto '{elemento['nombre']}' insertado en {celda}")

            # 2. Insertar imágenes/rangos (ya soporta múltiples celdas via _insertar_imagen)
            for elemento in self.config['elementos']:
                if elemento['tipo'] in ['imagen', 'rango'] and elemento['nombre'] in self.data['imagenes']:
                    self._insertar_imagen(wb_sid, elemento)

            # Guardar el resultado
            wb_sid.save(output_path)
            print(f"\n✅ SID generado correctamente en: {os.path.abspath(output_path)}")

        except Exception as e:
            print(f"\n❌ Error generando SID: {str(e)}")
            raise
        finally:
            app.quit()

    def _obtener_hoja(self, wb, sheet_identifier, book_type='sid'):
        """
        Obtiene una hoja por nombre o índice, con manejo de errores mejorado
        :param wb: Libro de trabajo (xlwings)
        :param sheet_identifier: Nombre o índice de la hoja
        :param book_type: 'sid' o 'tss' (para el mapeo de config)
        :return: Objeto hoja
        """
        try:
            # Si es string, buscar en la configuración
            sheet_index = self._obtener_hoja_indice(book_type, sheet_identifier)
            return wb.sheets[sheet_index]

        except Exception as e:
            available_sheets = "\n".join([f"- {s.name} (índice {i})" for i, s in enumerate(wb.sheets)])
            raise ValueError(
                f"No se pudo encontrar la hoja '{sheet_identifier}'.\n"
                f"Hojas disponibles:\n{available_sheets}"
            ) from e

    def _insertar_imagen(self, wb_sid, elemento):
        """Versión que soporta tamaño específico para imágenes y centrado en celda"""
        try:
            nombre = elemento['nombre']
            print(f"\n=== Insertando imagen '{nombre}' ===")

            # 1. Verificar existencia de la imagen
            img_path = os.path.abspath(self.data['imagenes'].get(nombre))
            if not os.path.exists(img_path):
                raise FileNotFoundError(
                    f"Imagen no encontrada.\nBuscada: {img_path}")

            # 2. Obtener hoja destino
            sheet_index = self._obtener_hoja_indice('sid', elemento['destino']['hoja'])
            sheet = wb_sid.sheets[sheet_index]
            print(f"Hoja destino: {sheet.name} (índice {sheet.index})")

            # 3. Obtener y convertir dimensiones (cm a puntos)
            width_cm = elemento['destino'].get('ancho')  # En cm
            height_cm = elemento['destino'].get('alto')  # En cm

            # Convertir cm a puntos (1 cm = 28.35 puntos)
            width = width_cm * 28.35 if width_cm is not None else None
            height = height_cm * 28.35 if height_cm is not None else None

            print(f"Configuración de tamaño - Ancho: {width_cm}cm ({width}pt), Alto: {height_cm}cm ({height}pt)")

            # 4. Procesar TODAS las celdas destino
            for celda in elemento['destino']['celdas']:
                try:
                    rango = sheet.range(celda)
                    print(f"Insertando en celda: {rango.address}")

                    # Calcular posición centrada
                    left = rango.left + 5
                    top = rango.top + 5

                    # Insertar imagen
                    picture = sheet.pictures.add(
                        img_path,
                        left=left,
                        top=top,
                        width=width,
                        height=height
                    )

                    # picture.api.ShapeRange.ZOrder(win32com.client.constants.msoSendToBack)
                    #
                    # # Opcional: Bloquear posición y tamaño
                    # picture.api.Placement = 1

                    # Mantener relación de aspecto si solo se especifica un dimensión
                    if width is not None and height is None:
                        # Mantener relación de aspecto basado en el ancho
                        img = Image.open(img_path)
                        aspect_ratio = img.height / img.width
                        picture.height = width * aspect_ratio
                        # Recalcular posición vertical después de ajustar altura
                        picture.top = rango.top + (rango.height - picture.height) / 2
                    elif height is not None and width is None:
                        # Mantener relación de aspecto basado en el alto
                        img = Image.open(img_path)
                        aspect_ratio = img.width / img.height
                        picture.width = height * aspect_ratio
                        # Recalcular posición horizontal después de ajustar ancho
                        picture.left = rango.left + (rango.width - picture.width) / 2

                    print(f"✅ Imagen insertada en {celda} - Tamaño: {width or 'auto'}x{height or 'auto'}")

                except Exception as e:
                    print(f"⚠️ Error insertando en {celda}: {type(e).__name__} - {str(e)}")

            return True

        except Exception as e:
            print(f"\n❌ ERROR insertando '{nombre}': {type(e).__name__}")
            print(f"Mensaje: {str(e)}")
            return False


# Uso del sistema
if __name__ == "__main__":
    processor = TSSProcessor('config.json')

    start_time = time.time()

    # Paso 1: Procesar TSS (extraer datos)
    tss_files = [f for f in os.listdir("TSS_PRUEBA") if f.endswith(('.xls', '.xlsx', '.xlsm'))]
    if not tss_files:
        print("❌ No se encontraron archivos TSS")
        exit()

    tss_path = os.path.join("TSS_PRUEBA", tss_files[0])
    processor.tss_path = tss_path

    nombre_sid = processor._generar_nombre_sid(tss_path)
    processor.procesar_tss(tss_path,nombre_sid)

    # Paso 2: Generar SID
    output_folder = "SIDs"
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, nombre_sid)

    processor.generar_sid(
        plantilla_path=processor.config['nombre_sid']['plantilla'],
        output_path=output_path
    )


    end_time = time.time()
    total_time = end_time - start_time
    print(f"\n⏱ Tiempo total de procesamiento: {total_time:.2f} segundos")
