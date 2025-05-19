import glob
import io
import os
import json
import re
import shutil
import traceback
import warnings
from collections import defaultdict
from datetime import timedelta

import win32com
from PIL import Image, ImageGrab
import openpyxl
import xlwings as xw
import win32com.client as win32
import win32gui
import win32con
import time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.reader.drawings')

# -*- coding: utf-8 -*-
def listar_ventanas_office():
    office_windows = []

    def callback(hwnd, _):
        title = win32gui.GetWindowText(hwnd)
        class_name = win32gui.GetClassName(hwnd)
        if (win32gui.IsWindowVisible(hwnd) and title and
                ("Excel" in title or "Office" in title or class_name in ['NUIDialog', '#32770'])):
            office_windows.append((hwnd, title, class_name))

    win32gui.EnumWindows(callback, None)
    return office_windows


def cerrar_dialogos_office():
    dialogs_closed = 0
    windows = listar_ventanas_office()
    for hwnd, title, class_name in windows:
        try:
            if class_name in ['NUIDialog', '#32770'] and "Excel" not in title:
                win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
                time.sleep(2)
                if win32gui.IsWindow(hwnd):
                    win32gui.SendMessage(hwnd, win32con.WM_SYSCOMMAND, win32con.SC_CLOSE, 0)
                    time.sleep(1)
                if not win32gui.IsWindow(hwnd):
                    dialogs_closed += 1
        except:
            pass
    return dialogs_closed



class TSSInstance:
    """Representa un archivo TSS individual con sus metadatos"""
    def __init__(self, file_path,config_path='config.json'):

        self.file_path = file_path
        self.name = "DEFAULT_NAME"
        self.id = "DEFAULT_ID"
        self.data = {'textos': {}, 'imagenes': {}}
        self.resultados_dir = ""
        self.config = _cargar_configuracion(config_path)
        self._extraer_metadatos()


    def _extraer_metadatos(self):
        """Extrae name/id al inicializar cada instancia usando la configuración"""
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            # Usar nombres de configuración en lugar de nombres directos de hojas
            self.name = self._leer_celda(wb, "informacion", "H8")
            self.id = self._leer_celda(wb, "informacion", "H7")
            wb.close()
        except Exception as e:
            print(f"⚠️ Error extrayendo metadatos de {self.file_path}: {str(e)}")
            self.name = f"ERROR_{os.path.basename(self.file_path)}"
            self.id = time.strftime('%Y%m%d%H%M%S')

    def _leer_celda(self, wb, sheet_config_name, celda):

        global sheet_index
        try:
            # Obtener el índice de la hoja desde la configuración
            sheet_index = self._obtener_hoja_indice('tss', sheet_config_name)

            # Obtener la hoja por índice
            sheet = wb.worksheets[sheet_index]

            # Leer y limpiar el valor de la celda
            valor = sheet[celda].value
            return str(valor).strip() if valor is not None else ""

        except KeyError as e:
            print(f"⚠️ Error: No se encontró la hoja '{sheet_config_name}' en la configuración")
            return ""
        except IndexError as e:
            print(f"⚠️ Error: Índice {sheet_index} no existe en el workbook para hoja '{sheet_config_name}'")
            return ""
        except Exception as e:
            print(f"⚠️ Error leyendo celda {celda} de hoja '{sheet_config_name}': {str(e)}")
            return ""

    def _obtener_hoja_indice(self, workbook_type, sheet_name):
        return self.config['hojas'][workbook_type][sheet_name]

def _limpiar_texto(texto):
    """Limpio texto para usar en nombres de archivos"""
    return ''.join(c for c in texto if c not in '\\/:*?"<>|').replace(" ", "_")


def _cargar_configuracion(config_path):
    with open(config_path, encoding='utf-8') as f:
        return json.load(f)

OFFSET_BUSQUEDA = 12

class TSSBatchProcessor:
    """Procesa múltiples archivos TSS en lote"""

    def __init__(self, config_path='config.json'):
        self.config = _cargar_configuracion(config_path)
        self.tss_instances = []  # Lista de objetos TSSInstance
        self.total_time = 0

    def procesar_lote(self, tss_folder="TSS"):

        """Procesa todos los TSS en un directorio con medición de tiempo"""
        print("\n" + "=" * 50)
        print(" INICIANDO PROCESAMIENTO POR LOTES ")
        print("=" * 50 + "\n")
        """Procesa todos los TSS en un directorio"""
        tss_files = self._encontrar_archivos_tss(tss_folder)
        total_files = len(tss_files)
        start_time_total = time.monotonic()

        for i, tss_path in enumerate(tss_files, 1):
            print(f"\n📂 Procesando archivo {i} de {total_files}")
            file_start_time = time.monotonic()

            tss_instance = TSSInstance(tss_path)
            self.tss_instances.append(tss_instance)
            self._procesar_individual(tss_instance)

            file_time = time.monotonic() - file_start_time
            self.total_time += file_time
            print(f"⏱️ Tiempo archivo: {timedelta(seconds=file_time)}")

            # Estimación del tiempo restante
            remaining_files = total_files - i
            avg_time = self.total_time / i
            estimated_remaining = avg_time * remaining_files
            print(f"⏳ Estimado restante: {timedelta(seconds=estimated_remaining)}")

        total_elapsed = time.monotonic() - start_time_total
        print("\n" + "=" * 50)
        print(" RESUMEN DE TIEMPOS ")
        print("=" * 50)
        print(f"📊 Total archivos procesados: {total_files}")
        print(f"⏱️ Tiempo total: {timedelta(seconds=total_elapsed)}")
        print(f"⏱️ Tiempo promedio por archivo: {timedelta(seconds=total_elapsed / total_files if total_files else 0)}")
        print("=" * 50 + "\n")

    def _encontrar_archivos_tss(self, folder_path):
        """Encuentra todos los archivos Excel (.xlsx) en el directorio especificado"""
        tss_files = []
        try:
            if not os.path.exists(folder_path):
                print(f"⚠️ El directorio {folder_path} no existe")
                return tss_files

            for filename in os.listdir(folder_path):
                if filename.lower().endswith(('.xls', '.xlsx', '.xlsm')) and not filename.startswith('~$'):
                    full_path = os.path.join(folder_path, filename)
                    tss_files.append(full_path)

            print(f"📁 Encontrados {len(tss_files)} archivos TSS en {folder_path}")
            return tss_files
        except Exception as e:
            print(f"❌ Error buscando archivos TSS: {str(e)}")
            return []

    def _procesar_individual(self, tss_instance):
        """Procesamiento completo para un TSS"""
        print(f"\n🔁 Procesando {tss_instance.name}_{tss_instance.id}")

        # 1. Configurar rutas
        output_dir = os.path.join("resultados", f"{tss_instance.name}_{tss_instance.id}")
        os.makedirs(output_dir, exist_ok=True)

        valores = {
            'name': tss_instance.name,
            'id': tss_instance.id
        }
        nombre_archivo = self.config['nombre_sid']['formato'].format(**valores)
        nombre_archivo = _limpiar_texto(nombre_archivo)

        output_folder = "SIDs"
        os.makedirs(output_folder, exist_ok=True)
        output_path = os.path.join(output_folder, nombre_archivo)

        tss_instance.resultados_dir = os.path.join("resultados", f"{tss_instance.name}_{tss_instance.id}")
        os.makedirs(tss_instance.resultados_dir, exist_ok=True)

        # 2. Procesar contenido (adaptar tus métodos actuales)
        self._extraer_datos(tss_instance)

        self.procesar_fotos_antenas(tss_instance)

        self._generar_sid(
            tss_instance,
            self.config['nombre_sid']['plantilla'],
            output_path
        )
        print(f"✅ Proceso completado para {tss_instance.file_path}")

    # Configuración y helpers básicos

    def _obtener_hoja_indice(self, workbook_type, sheet_name):
        return self.config['hojas'][workbook_type][sheet_name]



    #Capturar informacion del tss

    def _extraer_datos(self, tss_instance):
        """Procesa el TSS agrupando elementos por tipo para optimización"""
        print(f"\n=== EXTRAYENDO DATOS DE {tss_instance.name}_{tss_instance.id} ===")
        wb_tss = None
        try:
            wb_tss = openpyxl.load_workbook(tss_instance.file_path, data_only=True)

            # Organizar elementos por tipo para procesamiento eficiente
            elementos_por_tipo = {
                'rango': [],
                'imagen': [],
                'texto': []
            }

            for elemento in self.config['elementos']:
                elementos_por_tipo[elemento['tipo']].append(elemento)

            # Procesar textos primero (más rápido)
            for elemento in elementos_por_tipo['texto']:
                self._procesar_texto(wb_tss, tss_instance, elemento)

            # Procesar rangos (requiere Excel COM)
            if elementos_por_tipo['rango']:
                self._procesar_rangos_agrupados(wb_tss, tss_instance, elementos_por_tipo['rango'])

            # Procesar imágenes
            for elemento in elementos_por_tipo['imagen']:
                self._procesar_imagen(wb_tss, tss_instance, elemento)

            print(f"✅ Extracción completada para {tss_instance.name}_{tss_instance.id}")
            return True

        except Exception as e:
            print(f"❌ Error en extracción de datos: {str(e)}")
            return False
        finally:
            wb_tss.close()
    # Procesamiento interno del tss

    def _procesar_texto(self, wb_tss, tss_instance, elemento):
        """Procesa un elemento de texto y lo almacena en la instancia"""
        try:
            sheet_index = self._obtener_hoja_indice('tss', elemento['origen']['hoja'])
            sheet = wb_tss.worksheets[sheet_index]
            valor = sheet[elemento['origen']['celda']].value
            tss_instance.data['textos'][elemento['nombre']] = str(valor).strip() if valor else ""
            print(f"Texto '{elemento['nombre']}' extraído: {tss_instance.data['textos'][elemento['nombre']][:50]}...")
        except Exception as e:
            print(f"⚠️ Error procesando texto {elemento['nombre']}: {str(e)}")

    def _procesar_rangos_agrupados(self, wb_tss, tss_instance, elementos_rango):
        """Procesa múltiples rangos usando Excel COM"""
        try:

            rangos_dict = {
                elem['nombre']: {
                    'rango': elem['origen']['rango'],
                    'hoja': elem['origen']['hoja']
                }
                for elem in elementos_rango
            }

            resultados = self.capturar_multiples_rangos(tss_instance, rangos_dict)


                # 4. Almacenar rutas de imágenes válidas
            for nombre, ruta_imagen in resultados.items():
                if ruta_imagen and os.path.exists(ruta_imagen):
                    tss_instance.data['imagenes'][nombre] = ruta_imagen
                    print(f"✅ Rango '{nombre}' guardado en {ruta_imagen}")
                else:
                    print(f"⚠️ No se pudo capturar el rango '{nombre}' o la imagen no existe")

            return True

        except Exception as e:
            print(f"❌ Error en procesamiento de rangos agrupados: {str(e)}")
            return False

    def capturar_multiples_rangos(self, tss_instance, rangos_dict):
        """Captura rangos usando Excel COM y guarda en la carpeta de la instancia"""
        excel = None
        resultados = {}
        wb = None

        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = True
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(os.path.abspath(tss_instance.file_path))
            cerrar_dialogos_office()

            os.makedirs(tss_instance.resultados_dir, exist_ok=True)

            for nombre, config in rangos_dict.items():
                output_path = os.path.join(tss_instance.resultados_dir, f"{nombre}.png")
                print("outputpath resultados dir", output_path)
                sheet_index = self._obtener_hoja_indice('tss', config['hoja']) + 1
                sheet = wb.Sheets(sheet_index)

                for intento in range(3):
                    try:
                        sheet.Range(config['rango']).CopyPicture(Appearance=1, Format=2)
                        time.sleep(2)

                        img = ImageGrab.grabclipboard()
                        if img:
                            img.save(output_path)
                            resultados[nombre] = output_path
                            print(f"✅ {nombre} guardado en {output_path}")
                            break
                    except Exception as e:
                        print(f"⚠️ Intento {intento + 1} para {nombre}: {str(e)}")
                        time.sleep(1)
                else:
                    resultados[nombre] = None
                    print(f"❌ No se pudo capturar el rango '{nombre}'")

            return resultados


        except Exception as e:
            print(f"❌ Error crítico en captura de rangos: {str(e)}")
            return {}  # Retornar diccionario vacío en caso de error crítico

        finally:
            # Cerrar todo correctamente
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception as e:
                print(f"⚠️ Error cerrando libro: {str(e)}")
            try:
                if excel is not None:
                    excel.DisplayAlerts = False
                    excel.Quit()
            except Exception as e:
                print(f"⚠️ Error cerrando Excel: {str(e)}")
            # Liberar recursos COM
            del wb
            del excel

    def _procesar_imagen(self, wb_tss, tss_instance, elemento):
        """Busca imágenes mostrando el rango de celdas de búsqueda"""
        try:
            sheet_index = self._obtener_hoja_indice('tss', elemento['origen']['hoja'])
            sheet = wb_tss.worksheets[sheet_index]
            celda = sheet[elemento['origen']['celda']]

            # Determinar coordenadas de búsqueda
            merged_range = self._encontrar_rango_combinado(celda, sheet)  # Pasar sheet como parámetro
            min_row, max_row, min_col, max_col = self._obtener_rango_expandido(celda, merged_range)

            # Convertir coordenadas numéricas a formato de letra de columna (A, B, C...)
            col_letter_start = openpyxl.utils.get_column_letter(min_col)
            col_letter_end = openpyxl.utils.get_column_letter(max_col)

            print(f"🔍 Buscando imagen {elemento['nombre']} en rango: "
                  f"{col_letter_start}{min_row}:{col_letter_end}{max_row} "
                  f"(Columnas {min_col}-{max_col}, Filas {min_row}-{max_row})")

            # Verificar si la hoja tiene imágenes antes de intentar acceder
            if not hasattr(sheet, '_images'):
                print(f"⚠️ Hoja {sheet.title} no contiene imágenes")
                return None

            # Buscar imagen en el rango
            for img in sheet._images:
                img_top = img.anchor._from.row + 1
                img_left = img.anchor._from.col + 1

                if (min_row <= img_top <= max_row) and (min_col <= img_left <= max_col):
                    img_path = os.path.join(tss_instance.resultados_dir, f"{elemento['nombre']}.png")
                    os.makedirs(os.path.dirname(img_path), exist_ok=True)  # Asegurar que el directorio existe

                    image_bytes = img._data()
                    image = Image.open(io.BytesIO(image_bytes))
                    image.save(img_path)
                    tss_instance.data['imagenes'][elemento['nombre']] = img_path
                    print(f"✅ Imagen '{elemento['nombre']}' encontrada en posición: "
                          f"Columna {img_left}, Fila {img_top}")
                    return img_path

            print(f"⚠️ Imagen {elemento['nombre']} no encontrada en el rango especificado")
            return None

        except Exception as e:
            print(f"❌ Error al buscar imagen: {str(e)}")
            return None

    def _encontrar_rango_combinado(self, target_cell, sheet):
        """Encontrar rango combinado para la celda objetivo"""
        for merged_cell in sheet.merged_cells.ranges:  # Usar sheet en lugar de self.sheet_tss
            if target_cell.coordinate in merged_cell:
                print(f"\n ✅ Celda combinada encontrada: {merged_cell.coord}")
                return merged_cell
        print(f"\n ℹ️ Celda no está combinada")
        return None

    def _obtener_rango_expandido(self, target_cell, merged_range):
        """Obtener rango ampliado para búsqueda de imagen"""
        if merged_range:
            min_row, max_row = merged_range.min_row, merged_range.max_row
            min_col, max_col = merged_range.min_col, merged_range.max_col
        else:
            min_row = max_row = target_cell.row
            min_col = max_col = target_cell.column

        # Ampliar rango con márgenes
        return (
            max(1, min_row - 1),  # expanded_min_row
            max_row,              # expanded_max_row
            max(1, min_col - 1),  # expanded_min_col
            max_col               # expanded_max_col
        )

    #Generacion de sid

    def _generar_sid(self, tss_instance, plantilla_path, output_path):
        """Genera el SID con los datos extraídos, soportando múltiples celdas destino"""
        print("\n=== GENERANDO SID ===")
        app = xw.App(visible=False)

        try:
            wb_sid = app.books.open(plantilla_path)

            # 1. Insertar textos (ahora soporta múltiples celdas destino)
            for elemento in self.config['elementos']:
                if elemento['tipo'] == 'texto' and elemento['nombre'] in tss_instance.data['textos']:
                    sheet_index = self._obtener_hoja_indice('sid', elemento['destino']['hoja'])
                    sheet = wb_sid.sheets[sheet_index]
                    valor = tss_instance.data['textos'][elemento['nombre']]

                    # Insertar el mismo valor en todas las celdas especificadas
                    for celda in elemento['destino']['celdas']:
                        sheet[celda].value = valor
                        print(f"Texto '{elemento['nombre']}' insertado en {celda}")

            # 2. Insertar imágenes/rangos (ya soporta múltiples celdas via _insertar_imagen)
            for elemento in self.config['elementos']:
                if elemento['tipo'] in ['imagen', 'rango'] and elemento['nombre'] in tss_instance.data['imagenes']:
                    self._insertar_imagen(wb_sid,tss_instance, elemento)

            self._insertar_fotos_antenas(wb_sid, tss_instance)
            self._actualizar_titulos_antenas(wb_sid, tss_instance)
            self._actualizar_sectores_con_tecnologias(wb_sid, tss_instance)

            # Guardar el resultado
            wb_sid.save(output_path)
            print(f"\n✅ SID generado correctamente en: {os.path.abspath(output_path)}")

        except Exception as e:
            print(f"\n❌ Error generando SID: {str(e)}")
            raise
        finally:
            app.quit()

    def _obtener_hoja(self, wb, sheet_identifier, book_type='sid'):
        """
        Obtiene una hoja por nombre o índice, con manejo de errores mejorado
        :param wb: Libro de trabajo (xlwings)
        :param sheet_identifier: Nombre o índice de la hoja
        :param book_type: 'sid' o 'tss' (para el mapeo de config)
        :return: Objeto hoja
        """
        try:
            # Si es string, buscar en la configuración
            sheet_index = self._obtener_hoja_indice(book_type, sheet_identifier)
            return wb.sheets[sheet_index]

        except Exception as e:
            available_sheets = "\n".join([f"- {s.name} (índice {i})" for i, s in enumerate(wb.sheets)])
            raise ValueError(
                f"No se pudo encontrar la hoja '{sheet_identifier}'.\n"
                f"Hojas disponibles:\n{available_sheets}"
            ) from e

    def _insertar_imagen(self, wb_sid, tss_instance, elemento):
        """Versión que soporta tamaño específico para imágenes y centrado en celda"""

        nombre = elemento['nombre']

        try:
            print(f"\n=== Insertando imagen '{nombre}' ===")

            # 1. Verificar existencia de la imagen
            img_path = os.path.abspath(tss_instance.data['imagenes'].get(nombre))
            if not os.path.exists(img_path):
                raise FileNotFoundError(
                    f"Imagen no encontrada.\nBuscada: {img_path}")

            # 2. Obtener hoja destino
            sheet_index = self._obtener_hoja_indice('sid', elemento['destino']['hoja'])
            sheet = wb_sid.sheets[sheet_index]
            print(f"Hoja destino: {sheet.name} (índice {sheet.index})")

            # 3. Obtener y convertir dimensiones (cm a puntos)
            width_cm = elemento['destino'].get('ancho')  # En cm
            height_cm = elemento['destino'].get('alto')  # En cm

            # Convertir cm a puntos (1 cm = 28.35 puntos)
            width = width_cm * 28.35 if width_cm is not None else None
            height = height_cm * 28.35 if height_cm is not None else None

            print(f"Configuración de tamaño - Ancho: {width_cm}cm ({width}pt), Alto: {height_cm}cm ({height}pt)")

            # 4. Procesar TODAS las celdas destino
            for celda in elemento['destino']['celdas']:
                try:
                    rango = sheet.range(celda)
                    print(f"Insertando en celda: {rango.address}")

                    # Calcular posición centrada
                    left = rango.left + 5
                    top = rango.top + 70
                    # Insertar imagen
                    picture = sheet.pictures.add(
                        img_path,
                        left=left,
                        top=top,
                        width=width,
                        height=height
                    )


                    # # Mantener relación de aspecto si solo se especifica una dimensión
                    if width is not None and height is None:
                        # Mantener relación de aspecto basado en el ancho
                        img = Image.open(img_path)
                        aspect_ratio = img.height / img.width
                        picture.height = width * aspect_ratio
                        # Recalcular posición vertical después de ajustar altura
                        picture.top = rango.top + (rango.height - picture.height) / 2
                    elif height is not None and width is None:
                        # Mantener relación de aspecto basado en el alto
                        img = Image.open(img_path)
                        aspect_ratio = img.width / img.height
                        picture.width = height * aspect_ratio
                        # Recalcular posición horizontal después de ajustar ancho
                        picture.left = rango.left + (rango.width - picture.width) / 2

                    print(f"✅ Imagen insertada en {celda} - Tamaño: {width or 'auto'}x{height or 'auto'}")

                except Exception as e:
                    print(f"⚠️ Error insertando en {celda}: {type(e).__name__} - {str(e)}")

            return True

        except Exception as e:
            print(f"\n❌ ERROR insertando '{nombre}': {type(e).__name__}")
            print(f"Mensaje: {str(e)}")
            return False

    def obtener_tecnologias(nombre_archivo):
        if '(' in nombre_archivo and ')' in nombre_archivo:
            tech_part = nombre_archivo.split('(')[-1].split(')')[0]
            return tech_part.replace('-', ' + ')
        return ""

    def procesar_fotos_antenas(self, tss_instance):
        """Procesa las fotos de antenas para una instancia TSS"""
        try:
            print("\n=== PROCESANDO FOTOS DE ANTENAS ===")

            # Crear subcarpeta para este TSS si no existe
            proyecto_folder = os.path.join("resultados", f"{tss_instance.name}_{tss_instance.id}")
            os.makedirs(proyecto_folder, exist_ok=True)

            # Configuración de sectores y antenas
            lista_sectores = ['a', 'b', 'c']
            lista_antenas = [1, 2, 3, 4]

            self.buscar_antenas_por_sectores(
                tss_instance.file_path,
                lista_sectores,
                lista_antenas,
                proyecto_folder
            )

        except Exception as e:
            print(f"⚠️ Error procesando fotos de antenas: {str(e)}")

    def buscar_antenas_por_sectores(self, excel_path, lista_sectores, lista_antenas, output_folder):
        """Versión adaptada del método original"""
        global frase_busqueda
        wb = None
        try:
            wb = load_workbook(excel_path, data_only=True)

            # Usar índice de hoja desde configuración
            sheet_index = self._obtener_hoja_indice('tss', 'torres')
            sheet = wb.worksheets[sheet_index]

            imagenes_dict = {}
            for img in sheet._images:
                pos = img.anchor._from
                excel_row = pos.row + 1
                excel_col = pos.col + 1
                imagenes_dict[(excel_row, excel_col)] = img

            merged_ranges = list(sheet.merged_cells.ranges)

            # Crear carpetas Antena_X dentro del proyecto
            for antena in lista_antenas:
                folder_path = os.path.join(output_folder, f"Antena_{antena}")
                os.makedirs(folder_path, exist_ok=True)

            # Buscar todas las combinaciones
            for sector in lista_sectores:
                for antena in lista_antenas:
                    try:
                        frase_busqueda = f"foto general de la antena {antena} sector {sector}"
                        print(f"\nBuscando: {frase_busqueda}")

                        # Buscar celda con texto
                        target_cell = None
                        descripcion_tecnica = None

                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value and frase_busqueda in str(cell.value).lower():
                                    target_cell = cell
                                    break
                            if target_cell:
                                celda_encontrada = f"{get_column_letter(target_cell.column)}{target_cell.row}"
                                print(f"Texto encontrado en la celda: {celda_encontrada}")

                                # Extraer descripción técnica
                                texto_completo = str(target_cell.value)
                                if ":" in texto_completo:
                                    _, descripcion = texto_completo.split(":", 1)
                                    descripcion_tecnica = descripcion.strip()[:30]
                                    descripcion_tecnica = descripcion_tecnica.replace("/", "-").replace("\\", "-")
                                break

                        if not target_cell:
                            print(f"No encontrado: {frase_busqueda}")
                            continue

                        # Detectar celdas combinadas
                        merged_range = None
                        for merged in merged_ranges:
                            if (merged.min_row <= target_cell.row <= merged.max_row and
                                    merged.min_col <= target_cell.column <= merged.max_col):
                                merged_range = merged
                                break

                        # Definir rango de búsqueda
                        rango_filas = range(max(1, target_cell.row - OFFSET_BUSQUEDA), target_cell.row)
                        start_col = merged_range.min_col if merged_range else target_cell.column
                        end_col = merged_range.max_col if merged_range else target_cell.column
                        rango_columnas = range(start_col, end_col + 1)

                        # Buscar imagen en el diccionario
                        imagen_encontrada = False
                        for fila in rango_filas:
                            for col in rango_columnas:
                                if (fila, col) in imagenes_dict:
                                    img = imagenes_dict[(fila, col)]
                                    folder = os.path.join(output_folder, f"Antena_{antena}")

                                    # Nombre del archivo
                                    if descripcion_tecnica:
                                        filename = f"Antena_{antena}_Sector_{sector}_({descripcion_tecnica}).png"
                                    else:
                                        filename = f"Antena_{antena}_Sector_{sector}.png"
                                    output_path = os.path.join(folder, filename)

                                    try:
                                        img_data = img._data()
                                        with open(output_path, "wb") as f:
                                            f.write(img_data)

                                        # Verificar imagen
                                        with Image.open(output_path) as img_pil:
                                            img_pil.verify()

                                        print(f"Imagen guardada en: {output_path}")
                                        imagen_encontrada = True
                                        break

                                    except Exception as e:
                                        print(f"Error guardando imagen: {str(e)}")

                            if imagen_encontrada:
                                break

                        if not imagen_encontrada:
                            print(f"¡Imagen no encontrada en el rango especificado!")

                    except Exception as e:
                        print(f"Error procesando {frase_busqueda}: {str(e)}")
                        continue
        finally:
            if wb:
                wb.close()


    def _insertar_fotos_antenas(self, wb_sid, tss_instance):
        """Inserta las fotos de las antenas generando títulos individuales"""
        try:
            print("\n=== INSERTANDO FOTOS DE ANTENAS ===")
            resultados_dir = os.path.abspath(tss_instance.resultados_dir)

            if not os.path.exists(resultados_dir):
                print(f"❌ Carpeta no encontrada: {resultados_dir}")
                return {}

            # Obtener hoja de trabajo
            sheet = wb_sid.sheets[self._obtener_hoja_indice('sid', 'antenas')]

            # Configuración de imágenes
            width = 9 * 28.35  # 10 cm a puntos
            height = 14 * 28.35  # 15 cm a puntos

            posiciones_base = {
                1: ("B20", "G20", "L20"),
                2: ("C74", "H74", "M74"),
                3: ("C124", "H124", "M124"),
                4: ("C174", "H174", "M174")
            }

            titulos_antenas = {}
            antenas_con_sectores_diferentes = []

            for antena in range(1, 5):
                antena_folder = os.path.join(resultados_dir, f"Antena_{antena}")
                if not os.path.exists(antena_folder):
                    continue

                print(f"\n📡 Procesando Antena {antena}")
                tecnologias_por_sector = []
                tecnologias_totales = set()

                # Procesar cada sector
                for sector_idx, sector in enumerate(['a', 'b', 'c']):
                    celda = posiciones_base[antena][sector_idx]
                    patron = os.path.join(antena_folder, f"Antena_{antena}_Sector_{sector}*.png")

                    sector_tecnologias = set()

                    for img_path in glob.glob(patron):
                        # Extraer tecnologías del nombre de archivo
                        if '(' in img_path and ')' in img_path:
                            tech_part = img_path.split('(')[1].split(')')[0]
                            for tech in [t.strip() for t in tech_part.replace('-', ',').split(',') if t.strip()]:
                                sector_tecnologias.add(tech)
                                tecnologias_totales.add(tech)

                        # Insertar imagen
                        try:
                            with Image.open(img_path) as img:
                                img.verify()
                                rango = sheet.range(celda)
                                picture = sheet.pictures.add(
                                    img_path,
                                    left=rango.left + (rango.width - width) / 2,
                                    top=rango.top + (rango.height - height) / 2,
                                    width=width,
                                    height=height
                                )
                                print(f"✅ Insertada {os.path.basename(img_path)} en {celda}")
                        except Exception as e:
                            print(f"❌ Error con {img_path}: {str(e)}")
                    tecnologias_por_sector.append(sector_tecnologias)

                sectores_diferentes = False
                if len(tecnologias_por_sector) > 1:
                    primer_sector = tecnologias_por_sector[0]
                    for sector_tech in tecnologias_por_sector[1:]:
                        if sector_tech != primer_sector:
                            sectores_diferentes = True
                            break

                if sectores_diferentes:
                    print(f"🔴 Antena {antena} tiene sectores con tecnologías diferentes")
                    antenas_con_sectores_diferentes.append(antena)
                else:
                    print(f"🟢 Antena {antena} tiene sectores con las mismas tecnologías")

                # Generar título para esta antena
                if tecnologias_totales:
                    tech_list = sorted(tecnologias_totales)
                    titulo = " + ".join(tech_list[:-1] + [tech_list[-1]] if len(tech_list) > 1 else tech_list)
                    titulos_antenas[antena] = titulo[0] if isinstance(titulo, list) else titulo
                    print(f"🔹 Tecnologías Antena {antena}: {titulos_antenas[antena]}")
                else:
                    titulos_antenas[antena] = "Sin tecnologías"

            if antenas_con_sectores_diferentes:
                print("\n=== ANTENAS CON SECTORES DIFERENTES DETECTADAS ===")
                print(f"Antenas a actualizar: {', '.join(map(str, antenas_con_sectores_diferentes))}")
                self._actualizar_sectores_con_tecnologias(wb_sid, tss_instance)

            return titulos_antenas

        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return {}

    def _actualizar_sectores_con_tecnologias(self, wb_sid, tss_instance):
        """Actualiza los sectores con sus tecnologías específicas cuando son diferentes"""
        try:
            print("\n=== BUSCANDO SECTORES PARA ACTUALIZAR ===")
            sheet = wb_sid.sheets[self._obtener_hoja_indice('sid', 'antenas')]

            print("🔍 Buscando sectores en la hoja...")

            # Aquí iría el resto de la lógica para actualizar los sectores
            # que ya tenías implementada en tu método original

        except Exception as e:
            print(f"❌ Error al actualizar sectores: {str(e)}")

    def verificar_posicion_imagenes(sheet, celda_objetivo):
        """Muestra información de posición de todas las imágenes en la hoja"""
        print(f"\n🔍 Verificando imágenes en hoja '{sheet.name}':")

        for shape in sheet.shapes:
            if shape.type == 'picture':
                # Obtener posición y tamaño
                top = shape.top
                left = shape.left
                height = shape.height
                width = shape.width

                # Convertir posición a coordenadas de celda
                row = int(top / sheet.range('A1').height) + 1
                col = int(left / sheet.range('A1').width) + 1
                letra_col = openpyxl.utils.get_column_letter(col)

                print(f"\n📸 Imagen: {shape.name}")
                print(f"▸ Posición: {letra_col}{row}")
                print(f"▸ Tamaño: {width:.2f}x{height:.2f} puntos")
                print(f"▸ Celda objetivo: {celda_objetivo}")

                # Verificar si coincide con la celda objetivo
                celda_actual = f"{letra_col}{row}"
                if celda_actual == celda_objetivo:
                    print("✅ Coincide con la posición esperada")
                else:
                    print(f"⚠️ Desplazada! Diferencia: {abs(col - openpyxl.utils.column_index_from_string(celda_objetivo[0]))} columnas, "
                          f"{abs(row - int(celda_objetivo[1:]))} filas")
    def _actualizar_titulos_antenas(self, wb_sid, tss_instance):
        """Actualiza los títulos de antenas en el SID con:
        - Código de sitio (ID) en los TextBoxes con 'XXX'
        - Tecnologías en los TextBoxes con 'TECH{num}'

        Args:
            wb_sid: Workbook de Excel (SID)
            tss_instance: Instancia del TSS con los datos a actualizar
        """
        try:
            print("\n" + "="*50)
            print(" ACTUALIZANDO TÍTULOS DE ANTENAS ")
            print("="*50)

            # 1. Configuración inicial
            sheet = self._obtener_hoja_antenas(wb_sid)
            codigo_sitio = tss_instance.id.upper()

            # 2. Procesar cada grupo de antenas
            for grupo in self._obtener_grupos_antenas(sheet):
                self._procesar_grupo_antenas(grupo, codigo_sitio, tss_instance)

            print("\n" + "="*50)
            print(" RESUMEN DE ACTUALIZACIONES ")
            print("="*50)
            self._imprimir_resumen_actualizaciones(sheet)

        except Exception as e:
            print(f"\n❌ ERROR CRÍTICO: {str(e)}")
            traceback.print_exc()

    def _obtener_hoja_antenas(self, wb_sid):
        """Obtiene la hoja de antenas del SID"""
        try:
            indice = self._obtener_hoja_indice('sid', 'antenas')
            return wb_sid.sheets[indice]
        except Exception as e:
            raise Exception(f"No se pudo obtener la hoja de antenas: {str(e)}")

    def _obtener_grupos_antenas(self, sheet):
        """Generador que devuelve cada grupo de antenas en la hoja"""
        for shape in sheet.shapes:
            if shape.type == 'group':
                yield shape


    def _procesar_grupo_antenas(self, grupo, codigo_sitio, tss_instance):
        """Procesa un grupo de antenas identificando por TECH*N*"""
        try:
            # 1. Extraer número de antena del contenido TECH
            num_antena = self._extraer_num_antena(grupo)
            if num_antena is None:
                return

            # 2. Actualizar código en el TextBox con XXX
            for item in grupo.api.GroupItems:
                if item.Type == 17 and "XXX" in str(item.TextFrame2.TextRange.Text):
                    item.TextFrame2.TextRange.Text = item.TextFrame2.TextRange.Text.replace("XXX", codigo_sitio)
                    print(f"✓ Antena {num_antena}: Código actualizado")

            # 3. Actualizar tecnologías en el TextBox TECH*N*
            folder_antena = os.path.join(tss_instance.resultados_dir, f"Antena_{num_antena}")
            if os.path.exists(folder_antena):
                tecnologias = self._extraer_tecnologias(folder_antena)
                if tecnologias:
                    for item in grupo.api.GroupItems:
                        if item.Type == 17 and f"TECH{num_antena}" in str(item.TextFrame2.TextRange.Text):
                            self._actualizar_textbox_tecnologias(item, tecnologias)
                            print(f"✓ Antena {num_antena}: Tecnologías actualizadas -> {', '.join(tecnologias)}")

        except Exception as e:
            print(f"⚠️ Error procesando grupo: {str(e)}")

    def _extraer_tecnologias(self, folder_path):
        """Extrae tecnologías de nombres de archivo"""
        tecnologias = set()
        for filename in os.listdir(folder_path):
            if filename.endswith('.png') and '(' in filename:
                tech_part = filename.split('(')[1].split(')')[0]
                tecnologias.update(t.strip() for t in re.split(r'[, \-+]', tech_part) if t.strip())
        return sorted(tecnologias)

    def _actualizar_textbox_tecnologias(self, textbox, tecnologias):
        """Formatea el TextBox de tecnologías"""
        texto = " + ".join(tecnologias)
        text_range = textbox.TextFrame2.TextRange
        text_range.Text = texto
        text_range.Font.Fill.ForeColor.RGB = (160 << 16) | (75 << 8) | 1  # RGB(1, 75, 160) → BGR(160, 75, 1)
        text_range.Font.Bold = True
    def _extraer_num_antena(self, grupo):
        """Extrae el número de antena del TextBox TECH"""
        for item in grupo.api.GroupItems:
            if item.Type == 17 and "TECH" in str(item.TextFrame2.TextRange.Text):
                try:
                    return int(str(item.TextFrame2.TextRange.Text).replace("TECH", ""))
                except:
                    return None
        return None
    def _clasificar_componentes(self, grupo):
        """Clasifica los TextBoxes del grupo en códigos y tecnologías"""
        componentes = {
            'codigos': [],
            'tecnologias': []
        }

        for item in grupo.api.GroupItems:
            if item.Type == 17:  # Es TextBox
                texto = item.TextFrame2.TextRange.Text.strip()
                print("Se pillo textbox "+item)
                if "XXX" in texto:
                    componentes['codigos'].append(item)
                elif "TECH" in texto:
                    componentes['tecnologias'].append(item)

        return componentes

    def _actualizar_codigos(self, textboxes, codigo):
        """Actualiza todos los TextBoxes de código con el ID del sitio"""
        for tb in textboxes:
            try:
                texto_original = tb.TextFrame2.TextRange.Text
                nuevo_texto = texto_original.replace("XXX", codigo)
                tb.TextFrame2.TextRange.Text = nuevo_texto
                print(f"✓ Código actualizado en {tb.Name}")
            except Exception as e:
                print(f"⚠️ Error actualizando código en {tb.Name}: {str(e)}")

    def _actualizar_tecnologias(self, textboxes, tss_instance):
        """Actualiza los TextBoxes de tecnología con datos de las fotos"""
        for tb in textboxes:
            try:
                # Extraer número de antena del nombre (ej: "TECH1" -> 1)
                num_antena = int(re.search(r"TECH(\d+)", tb.Name).group(1))

                # Obtener tecnologías de las fotos
                folder_antena = os.path.join(tss_instance.resultados_dir, f"Antena_{num_antena}")
                tecnologias = self._extraer_tecnologias(folder_antena)

                if tecnologias:
                    self._aplicar_formato_tecnologias(tb, tecnologias)
                    print(f"✓ Tecnologías actualizadas en Antena {num_antena}: {', '.join(tecnologias)}")
                else:
                    print(f"⚠️ Antena {num_antena}: No se encontraron tecnologías")

            except Exception as e:
                print(f"⚠️ Error actualizando tecnologías en {tb.Name}: {str(e)}")

    def _extraer_tecnologias(self, folder_path):
        """Extrae tecnologías únicas de los nombres de archivo PNG"""
        tecnologias = set()

        if os.path.exists(folder_path):
            for filename in os.listdir(folder_path):
                if filename.endswith('.png') and '(' in filename:
                    # Extraer texto entre paréntesis
                    tech_str = filename.split('(')[1].split(')')[0]
                    # Dividir por separadores comunes
                    for tech in re.split(r'[,\-+]', tech_str):
                        tech_limpia = tech.strip()
                        if tech_limpia:
                            tecnologias.add(tech_limpia)

        return sorted(tecnologias)  # Orden alfabético

    def _aplicar_formato_tecnologias(self, textbox, tecnologias):
        """Aplica formato al TextBox de tecnologías"""
        texto = " + ".join(tecnologias)
        text_range = textbox.TextFrame2.TextRange
        text_range.Text = texto

        # Formato azul corporativo
        # text_range.Font.Fill.ForeColor.RGB = 0x0170c0
        # text_range.Font.Bold = True
        # text_range.Font.Size = 10  # Tamaño consistente

    def _imprimir_resumen_actualizaciones(self, sheet):
        """Muestra resumen de cambios realizados"""
        for shape in sheet.shapes:
            if shape.type == 'group':
                print(f"\nGrupo: {shape.name}")
                for item in shape.api.GroupItems:
                    if item.Type == 17:  # TextBox
                        texto = item.TextFrame2.TextRange.Text.strip()
                        print(f"  ├─ {item.Name}: {texto[:50]}...")
    def _extraer_tecnologias_de_fotos(self, folder_path):
        """Extrae tecnologías de los nombres de archivo en la carpeta de antena"""
        tecnologias = set()
        if os.path.exists(folder_path):
            for filename in os.listdir(folder_path):
                if filename.endswith('.png') and '(' in filename and ')' in filename:
                    tech_part = filename.split('(')[1].split(')')[0]
                    tecnologias.update(t.strip() for t in tech_part.replace('-', ',').split(',') if t.strip())
        return tecnologias
    # Métodos auxiliares (ya existentes)
    def _obtener_textboxes_antenas(self, sheet):
        """Recopila TextBox de antenas en grupos (8.1, 8.2, etc.)"""
        textboxes = {}
        for shape in sheet.shapes:
            if shape.type == 'group':
                try:
                    for sub_shape in shape.api.GroupItems:
                        if sub_shape.Type == 17:  # TextBox
                            texto = sub_shape.TextFrame2.TextRange.Text
                            match = re.search(r"8\.(\d+)", texto)
                            if match:
                                textboxes[int(match.group(1))] = sub_shape
                except Exception as e:
                    print(f"⚠️ Error en grupo {shape.name}: {str(e)}")
        return textboxes

    def _imprimir_textboxes_actualizados(self, sheet):
        """Imprime el contenido de todos los TextBox tipo 17 en grupos"""
        for shape in sheet.shapes:
            if shape.type == 'group':
                try:
                    print(f"\n🔍 Grupo: {shape.name}")
                    for i, sub_shape in enumerate(shape.api.GroupItems):
                        if sub_shape.Type == 17:  # TextBox
                            texto = sub_shape.TextFrame2.TextRange.Text.strip()
                            print(f"  📝 TextBox {i + 1}:")
                            print(f"     {texto}")
                except Exception as e:
                    print(f"⚠️ Error leyendo grupo {shape.name}: {str(e)}")

    def _actualizar_sectores_con_tecnologias(self, wb_sid, tss_instance):
        """Actualiza los TextBox tipo 1 (sectores) con tecnologías correspondientes"""
        try:
            print("\n=== ACTUALIZANDO SECTORES CON TECNOLOGÍAS ===")
            sheet = wb_sid.sheets[self._obtener_hoja_indice('sid', 'antenas')]

            # Procesar cada grupo en la hoja
            for shape in sheet.shapes:
                if shape.type == 'group':
                    try:
                        # Buscar TextBox de sectores (tipo 1) en el grupo
                        sectores = []
                        for sub_shape in shape.api.GroupItems:
                            if sub_shape.Type == 1:  # TextBox tipo 1
                                texto = sub_shape.TextFrame2.TextRange.Text.strip()
                                if texto.startswith("SECTOR"):
                                    sectores.append(sub_shape)

                        # Si encontramos los 3 sectores
                        if len(sectores) == 3:
                            # Extraer número de antena del grupo (ej: "Group 10" -> antena 1)
                            try:
                                antena_num = int(re.search(r'\d+', shape.name).group()) % 10
                                if antena_num == 0:
                                    antena_num = 10
                            except:
                                continue

                            # Procesar cada sector
                            for i, sector_shape in enumerate(sectores, 1):
                                sector_folder = os.path.join(tss_instance.resultados_dir, f"Antena_{antena_num}")
                                patron = os.path.join(sector_folder, f"Antena_{antena_num}_Sector_{i}*.png")

                                # Extraer tecnologías de los archivos
                                tecnologias = set()
                                for img_path in glob.glob(patron):
                                    if '(' in img_path and ')' in img_path:
                                        tech_part = img_path.split('(')[1].split(')')[0]
                                        tecnologias.update(
                                            t.strip() for t in tech_part.replace('-', ',').split(',') if t.strip())

                                # Actualizar texto del sector
                                if tecnologias:
                                    texto_original = sector_shape.TextFrame2.TextRange.Text.strip()
                                    nuevo_texto = f"{texto_original}\n({' + '.join(sorted(tecnologias))})"
                                    sector_shape.TextFrame2.TextRange.Text = nuevo_texto
                                    print(f"✅ Antena {antena_num} - Sector {i}: {nuevo_texto}")

                    except Exception as e:
                        print(f"⚠️ Error procesando grupo {shape.name}: {str(e)}")

        except Exception as e:
            print(f"❌ Error general: {str(e)}")
# Uso del sistema
if __name__ == "__main__":
    processor = TSSBatchProcessor('config.json')

    # Procesar todos los TSS encontrados
    processor.procesar_lote("TSS")

    # Alternativa para procesar uno específico
    # tss_instance = TSSInstance("ruta/especifica.xlsx")
    # processor._procesar_individual(tss_instance)
